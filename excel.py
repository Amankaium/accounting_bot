from time import sleep
from json import loads
from requests import get
from openpyxl import load_workbook, Workbook
from mytoken import token

def send_message(token, chat_id, message):
    get(f"https://api.telegram.org/bot{token}/sendMessage?chat_id={chat_id}&text={message}")


def add_expeses(user_id, message_text):
    try:
        user_excel = load_workbook(f"excel/{user_id}.xlsx")
        user_ws = user_excel[user_excel.sheetnames[0]]
    except:
        user_excel = Workbook()
        user_excel.save(f"excel/{user_id}.xlsx")
        user_ws = user_excel.active

    try:
        money = int(message_text.split()[0])
        description = message_text[len(str(money)) + 1 : ]
        user_ws.append([money, description])
        user_excel.save(f"excel/{user_id}.xlsx")
        message = "Расход записан"
    except:
        message = "Расход НЕ записан, проверьте формат"
    
    return message


def add_message(message_id, message_text, user_id):
    ws.append([message_id, message_text, user_id])
    excel_file.save("messages.xlsx")


while True:
    sleep(1)
    r = get(f"https://api.telegram.org/bot{token}/getUpdates")
    d = loads(r.text)

    for update in d["result"]:
        message_id = update["message"]["message_id"]
        excel_file = load_workbook("messages.xlsx")
        ws = excel_file[excel_file.sheetnames[0]]

        answered_lst = [cell.value for cell in ws["A"]]

        if message_id not in answered_lst:
            message_text = update["message"]["text"]
            user_id = update["message"]["from"]["id"]

            add_message(message_id, message_text, user_id)
            
            answer_message = add_expeses(user_id, message_text)

            send_message(token, update["message"]["chat"]["id"], answer_message)
        
        
f.close()
